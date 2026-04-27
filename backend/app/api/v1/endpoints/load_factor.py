"""
부하율 (LoadFactor) API 라우터

공정 그룹별 부하율 데이터 조회 엔드포인트
"""

import io
import logging
from datetime import datetime
from typing import Any
from urllib.parse import quote

import httpx
from app.api.v1.endpoints.aps_proxy import router as aps_proxy_router
from app.core.config import settings
from app.schemas.load_factor import (
    LoadFactorDetailRequest,
    LoadFactorGroupRequest,
    LoadFactorMainRequest,
)
from app.services.load_factor import LoadFactorService, get_load_factor_service
from fastapi import APIRouter, Body, Depends, Path, Query, Request
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

router = APIRouter()
# APS C# 프록시 하위 경로 공유 — /proxy/load-factor-detail 등이 여기서 노출됨.
router.include_router(aps_proxy_router)
logger = logging.getLogger(__name__)


def _error_response(status_code: int, label: str, exc: Exception) -> JSONResponse:
    logger.exception(f"[{label}] {exc}")
    detail = str(exc) if settings.DEBUG else None
    return JSONResponse(
        status_code=status_code,
        content={
            "success": False,
            "error": f"{label} 중 오류가 발생했습니다.",
            **({"detail": detail} if detail else {}),
        },
    )


@router.get("/oper-groups")
async def get_oper_groups(
    project_id: str = Path(..., description="프로젝트 ID"),
    plan_ver: str = Query(..., alias="planVer", description="Plan version"),
    service: LoadFactorService = Depends(get_load_factor_service),
):
    """공정 그룹 목록 조회"""
    try:
        data = await service.get_oper_groups(project_id, plan_ver)
        return {"success": True, "count": len(data), "data": data}
    except Exception as e:
        return _error_response(500, "공정 그룹 목록 조회", e)


@router.post("/main")
async def get_main(
    params: LoadFactorMainRequest,
    project_id: str = Path(..., description="프로젝트 ID"),
    service: LoadFactorService = Depends(get_load_factor_service),
):
    """공정 그룹별 총 부하율 조회"""
    try:
        return await service.get_main(project_id, params)
    except ValueError as e:
        return _error_response(400, "부하율 메인 조회", e)
    except Exception as e:
        return _error_response(500, "부하율 메인 조회", e)


@router.post("/group")
async def get_group(
    params: LoadFactorGroupRequest,
    project_id: str = Path(..., description="프로젝트 ID"),
    service: LoadFactorService = Depends(get_load_factor_service),
):
    """공정 그룹별 일간 item_group 분해 조회"""
    try:
        return await service.get_group(project_id, params)
    except ValueError as e:
        return _error_response(400, "부하율 그룹 조회", e)
    except Exception as e:
        return _error_response(500, "부하율 그룹 조회", e)


@router.post("/detail")
async def get_detail(
    params: LoadFactorDetailRequest,
    project_id: str = Path(..., description="프로젝트 ID"),
    service: LoadFactorService = Depends(get_load_factor_service),
):
    """공정 그룹별 수요별 상세 조회"""
    try:
        return await service.get_detail(project_id, params)
    except ValueError as e:
        return _error_response(400, "부하율 상세 조회", e)
    except Exception as e:
        return _error_response(500, "부하율 상세 조회", e)


# ──────────────────────────────────────────────────────────────────────────
# 세부데이터 Excel 다운로드 — 원본 APS excel 큐 서비스의 축약판.
#   원본 플로우:
#     프론트 downloadBigData → ExcelQueue 백엔드 → C# APS JSON → xlsx 생성 → 스트리밍 → saveAs
#   포팅 플로우 (동일 철학):
#     프론트 downloadBigData → 이 엔드포인트(POST) → C# APS JSON → openpyxl → StreamingResponse → saveAs
# ──────────────────────────────────────────────────────────────────────────
@router.post("/detail-excel")
async def get_detail_excel(
    request: Request,
    project_id: str = Path(..., description="프로젝트 ID"),
    body: dict[str, Any] = Body(
        ...,
        examples=[
            {
                "file_name": "LoadFactorByOperGroup_Detail_Export",
                "column_map": {"oper_group_id": "공정그룹ID", "str_date": "일자"},
                "data_parameter": {"planVer": "20260420-M-01", "operGroupIDs": []},
            }
        ],
    ),
):
    """공정그룹별 부하율 세부 데이터를 xlsx 로 다운로드.

    1. body.data_parameter 를 C# `RarStrByOperGroup/Detail` 로 forward (JSON)
    2. column_map 기준으로 헤더/컬럼 순서 구성
    3. openpyxl 로 xlsx 메모리 생성
    4. StreamingResponse 로 브라우저 다운로드 응답
    """
    column_map: dict[str, str] = body.get("column_map") or {}
    data_parameter: dict[str, Any] = body.get("data_parameter") or {}
    file_name: str = body.get("file_name") or "export"

    aps_url = (
        f"{settings.APS_BACKEND_BASE_URL}/api/aps/backend/"
        f"{project_id}/RarStrByOperGroup/Detail"
    )
    # 세션/인증 헤더는 원본 요청 것 그대로 forward. (aps_proxy.proxy_to_aps 와 동일 규칙)
    forward_headers: dict[str, str] = {}
    for key in ("cookie", "authorization"):
        if key in request.headers:
            forward_headers[key] = request.headers[key]
    forward_headers.setdefault("content-type", "application/json")

    try:
        async with httpx.AsyncClient(timeout=settings.QUERY_TIMEOUT_SECONDS) as client:
            resp = await client.post(
                aps_url, json=data_parameter, headers=forward_headers
            )
        if resp.status_code == 401:
            return JSONResponse(
                status_code=401,
                content={
                    "success": False,
                    "error": "APS 인증이 필요합니다. 로그인 후 다시 시도하세요.",
                },
            )
        if resp.status_code != 200:
            logger.error(
                f"[LoadFactor detail-excel] C# 응답 오류 "
                f"status={resp.status_code}, body={resp.text[:200]}"
            )
            return _error_response(
                502, "APS 백엔드 호출", Exception(f"status={resp.status_code}")
            )
        payload = resp.json()
    except httpx.ConnectError as e:
        logger.error(f"[LoadFactor detail-excel] 연결 실패: {aps_url} ({e})")
        return _error_response(502, "APS 백엔드 연결", e)
    except Exception as e:
        return _error_response(500, "APS 백엔드 호출", e)

    # JSendResponse: { status, data: [...], ... } 또는 바로 [...] 일 수 있음.
    rows: list[dict[str, Any]] = payload.get("data") if isinstance(payload, dict) else []
    if not isinstance(rows, list):
        rows = []

    # 컬럼 순서/헤더 결정
    if column_map:
        bindings = list(column_map.keys())
        headers = [column_map[b] for b in bindings]
    elif rows:
        bindings = list(rows[0].keys())
        headers = bindings
    else:
        bindings = []
        headers = []

    # xlsx 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "Detail"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4568E0")

    # 헤더
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill

    # 데이터
    for row_idx, item in enumerate(rows, start=2):
        for col_idx, binding in enumerate(bindings, start=1):
            ws.cell(row=row_idx, column=col_idx, value=item.get(binding, ""))

    # 기본 컬럼 너비
    for col_idx in range(1, len(bindings) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = datetime.now().strftime("%Y%m%d")
    filename = f"{file_name}_{today}.xlsx"
    # RFC 5987 encoding for 한글 파일명 safety
    quoted = quote(filename)
    content_disposition = (
        f"attachment; filename=\"{filename}\"; filename*=UTF-8''{quoted}"
    )

    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": content_disposition},
    )
